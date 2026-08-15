"""Die Datensätze, aus denen dieser Server liest — namentlich gepinnt.

Warum nicht mehr über einen Organisationsfilter: Bis 2026-08-14 filterte jede
CKAN-Suche auf ``organization:staatssekretariat-fur-wirtschaft-seco``. Diese
Organisation gibt es auf opendata.swiss nicht (mehr) — ``organization_show``
antwortet 404, und in den 176 Einträgen von ``organization_list`` kommt kein
SECO vor. Jede Suche lieferte deshalb null Treffer, und die Werkzeuge
antworteten «Keine SECO-Datensätze gefunden»: ein Namensabgleich, der ins Leere
läuft, sieht genau aus wie eine leere Suche.

Ein Literal-Register statt eines Namensabgleichs ist die Antwort darauf —
dasselbe Muster wie ``CANTON_INSTITUTION_IDS`` in ``swiss-procurement-mcp``.
Es verschiebt den Fehler von *still* nach *laut*: ``test_live.py`` prüft jede
Kennung gegen die Quelle, und eine verschwundene Kennung ist dann ein roter
Test statt einer leeren Antwort.

**Herausgeber ist das BFS, Datenquelle ist das SECO.** Die Tabelle
``je-d-03.03.00.01`` führt drei Reihen nebeneinander und benennt beide Quellen
in ihrer Fusszeile: «Quelle: BFS – Erwerbslosenstatistik» und «Quelle:
Staatssekretariat für Wirtschaft (SECO)». Die registrierten Arbeitslosen und
Stellensuchenden sind SECOs eigene Zahlen aus dem RAV-System; das BFS
veröffentlicht sie. Der Server liest also weiterhin SECO-Zahlen, nur über einen
anderen Weg.

**Die drei Reihen sind nicht austauschbar.** Sie stehen im selben Blatt
untereinander und sehen dadurch vergleichbar aus:

===============================================  ======  ======
Reihe                                              2000    2025
===============================================  ======  ======
Registrierte Stellensuchende gemäss SECO          124.6   214.1
Registrierte Arbeitslose gemäss SECO               72.0   133.7
Erwerbslose (ILO) des BFS                         126.5   248.5
===============================================  ======  ======

Im Jahr 2000 ist die ILO-Zahl das **1.76-fache** der registrierten. Wer die
eine für die andere einsetzt, weil gerade nur die eine erreichbar ist,
produziert eine Zahl, die plausibel aussieht und um drei Viertel danebenliegt.
``REIHEN`` hält die Beschriftungen deshalb wörtlich, statt sie über die
Zeilenposition zu erraten.
"""

from __future__ import annotations

import io
from dataclasses import dataclass
from typing import Any

CKAN_BASE = "https://opendata.swiss/api/3/action"


@dataclass(frozen=True)
class Datensatz:
    """Ein gepinnter Datensatz auf opendata.swiss.

    ``ckan_id`` ist die UUID und nicht der Slug: Slugs tragen bei
    Namenskollisionen eine angehängte Ziffer (hier ``…stellensuchende4``), die
    sich beim nächsten Harvest verschieben kann. Der Slug steht trotzdem dabei,
    weil eine UUID allein niemandem sagt, worum es geht.
    """

    ckan_id: str
    slug: str
    titel: str
    herausgeber: str
    datenquelle: str
    blatt: str
    einheit: str

    @property
    def portal_url(self) -> str:
        return f"https://opendata.swiss/de/dataset/{self.slug}"


# Die Jahresreihe. Einzige zurzeit maschinenlesbare Quelle für die
# registrierten Arbeitslosen und Stellensuchenden des SECO — geprüft am
# 2026-08-14 gegen den ganzen Bestand von opendata.swiss.
JAHRESREIHE = Datensatz(
    ckan_id="13f60916-3df1-495a-9b30-4e9b1daea562",
    slug="erwerbslose-gemass-ilo-registrierte-arbeitslose-und-registrierte-stellensuchende4",
    titel="Erwerbslose gemäss ILO, registrierte Arbeitslose und registrierte Stellensuchende",
    herausgeber="Bundesamt für Statistik (BFS)",
    datenquelle="Staatssekretariat für Wirtschaft (SECO)",
    blatt="T3.3.0.1",
    einheit="Tausend Personen, Jahresdurchschnitt",
)

GEPINNTE_DATENSAETZE: tuple[Datensatz, ...] = (JAHRESREIHE,)

# Die Zeilenbeschriftungen, wörtlich aus dem Blatt. Der Parser sucht sie als
# Präfix, statt auf Zeilennummern zu zeigen: schiebt das BFS eine Zeile ein,
# ist eine falsche Zahl das Ergebnis einer Positionsannahme, aber ein leeres
# Ergebnis das einer Beschriftungssuche. Leer ist das bessere Scheitern.
REIHEN: dict[str, str] = {
    "registrierte_arbeitslose": "Registrierte Arbeitslose gemäss SECO",
    "registrierte_stellensuchende": "Registrierte Stellensuchende gemäss SECO",
    "erwerbslose_ilo": "Erwerbslose (ILO) des BFS",
}


class TabelleNichtLesbarError(RuntimeError):
    """Die Arbeitsmappe kam an, trug aber nicht die erwartete Form."""


def parse_jahresreihe(payload: bytes) -> dict[str, Any]:
    """Liest die drei Jahresreihen aus der BFS-Arbeitsmappe.

    Liefert ``{"years": [...], "series": {schlüssel: {jahr: wert}}, "labels":
    {...}}``. Fehlt eine erwartete Reihe, fliegt ``TabelleNichtLesbarError`` mit
    den tatsächlich gefundenen Beschriftungen in der Meldung — eine
    Formänderung soll benannt werden und nicht als leere Reihe durchgehen.

    Import lokal, damit ein fehlendes ``openpyxl`` erst beim Aufruf auffällt
    und nicht schon beim Import des Servers; dieselbe Regel wie bei ``pypdf``
    in ``uvg.py``.
    """
    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 — jede Lesefehlerart meint dasselbe
        raise TabelleNichtLesbarError(f"Arbeitsmappe nicht lesbar: {exc}") from exc

    if JAHRESREIHE.blatt not in wb.sheetnames:
        raise TabelleNichtLesbarError(
            f"Blatt {JAHRESREIHE.blatt!r} fehlt. Vorhanden: {wb.sheetnames}"
        )
    zeilen = [list(r) for r in wb[JAHRESREIHE.blatt].iter_rows(max_row=12, values_only=True)]

    jahre: list[int] = []
    for zeile in zeilen:
        kandidaten = [c for c in zeile[1:] if isinstance(c, int) and 1990 <= c <= 2100]
        if len(kandidaten) >= 5:
            jahre = kandidaten
            break
    if not jahre:
        raise TabelleNichtLesbarError("keine Jahres-Kopfzeile gefunden")

    beschriftungen = [str(z[0]).strip() for z in zeilen if isinstance(z[0], str)]
    reihen: dict[str, dict[int, float]] = {}
    gefunden: dict[str, str] = {}
    for schluessel, praefix in REIHEN.items():
        for zeile in zeilen:
            if not isinstance(zeile[0], str) or not zeile[0].strip().startswith(praefix):
                continue
            werte = [w for w in zeile[1:] if isinstance(w, int | float)]
            reihen[schluessel] = dict(zip(jahre, werte, strict=False))
            gefunden[schluessel] = zeile[0].strip()
            break

    fehlend = sorted(set(REIHEN) - set(reihen))
    if fehlend:
        raise TabelleNichtLesbarError(
            f"Reihen nicht gefunden: {[REIHEN[f] for f in fehlend]}. "
            f"Vorhandene Beschriftungen: {beschriftungen}"
        )
    return {"years": jahre, "series": reihen, "labels": gefunden}


def herkunftszeile(datensatz: Datensatz = JAHRESREIHE) -> str:
    """Die Attributionszeile, die jede Antwort mitführt.

    Nennt beide Häuser. Der Server behauptet damit nicht mehr, SECOs eigenes
    Portal zu lesen — es hat keine Schnittstelle, an der das ginge.
    """
    return (
        f"Datenquelle: {datensatz.datenquelle}, veröffentlicht durch das "
        f"{datensatz.herausgeber} (Tabelle `{datensatz.blatt}`, "
        f"{datensatz.portal_url})"
    )
